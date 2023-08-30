import pandas as pd
from datetime import date
from dbfread import DBF
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Ruta a los archivos DBF
pedidoc_files = {
    'LAGUNA': 'LAGUNA/pedidoc.DBF',
    'MTY': 'MTY/pedidoc.DBF'
}

pedidod_files = {
    'LAGUNA': 'LAGUNA/PEDIDOD.DBF',
    'MTY': 'MTY/PEDIDOD.DBF'
}

clientes_files = {
    'LAGUNA': 'LAGUNA/clientes.DBF',
    'MTY': 'MTY/clientes.DBF'
}

# Definir fecha de inicio y fecha de fin
fecha_inicio = date(2023, 8, 14)
fecha_fin = date(2023, 8, 26)

# Crear el archivo de Excel
archivo_excel = f'results/MTY_LAGUNA_del_{fecha_inicio.strftime("%Y-%m-%d")}_al_{fecha_fin.strftime("%Y-%m-%d")}.xlsx'
workbook = Workbook()

# Crear una hoja para los NO_PED agrupados
sheet_agregado = workbook.create_sheet('NO_PED_AGRUPADOS')
sheet_agregado.append(['CIUDAD', 'NO_PED'])

for ciudad in pedidoc_files.keys():
    # Leer el archivo pedidoc DBF
    pedidoc_dbf = DBF(pedidoc_files[ciudad], encoding='latin1')

    # Crear DataFrame de Pandas a partir de pedidoc
    df_pedidoc = pd.DataFrame(iter(pedidoc_dbf))

    # Filtrar pedidos con TOTAL_PED < 3480 y fecha entre fecha_inicio y fecha_fin
    df_filtrado1 = df_pedidoc[
        (df_pedidoc['TOTAL_PED'] < 3480) &
        ((df_pedidoc['F_ALTA_PED'] >= fecha_inicio) & (df_pedidoc['F_ALTA_PED'] <= fecha_fin) & (df_pedidoc['STATUS'] == 'Surtido') | (df_pedidoc['STATUS'] == 'Por Surtir'))

    ]

    # Filtrar pedidos con TOTAL_PED < 6960 y fecha entre fecha_inicio y fecha_fin
    df_filtrado2 = df_pedidoc[
        (df_pedidoc['TOTAL_PED'] < 6960) &
        ((df_pedidoc['F_ALTA_PED'] >= fecha_inicio) & (df_pedidoc['F_ALTA_PED'] <= fecha_fin) & (df_pedidoc['STATUS'] == 'Surtido') | (df_pedidoc['STATUS'] == 'Por Surtir'))

    ]
    # Filtrar pedidos con fecha entre fecha_inicio y fecha_fin
    df_filtrado3 = df_pedidoc[
        ((df_pedidoc['F_ALTA_PED'] >= fecha_inicio) & (df_pedidoc['F_ALTA_PED'] <= fecha_fin) & (df_pedidoc['STATUS'] == 'Surtido') | (df_pedidoc['STATUS'] == 'Por Surtir'))
    ]

    # Leer el archivo pedidod DBF
    pedidod_dbf = DBF(pedidod_files[ciudad], encoding='latin1')

    # Crear DataFrame de Pandas a partir de pedidod
    df_pedidod = pd.DataFrame(iter(pedidod_dbf))

    # Leer el archivo clientes DBF
    clientes_dbf = DBF(clientes_files[ciudad], encoding='latin1')

    # Crear DataFrame de Pandas a partir de clientes
    df_clientes = pd.DataFrame(iter(clientes_dbf))

    # Filtrar clientes con CVE_ZON == 2
    df_clientes_filtrado = df_clientes[df_clientes['CVE_ZON'] != 2]
    df_clientes_filtrado_local = df_clientes[df_clientes['CVE_ZON'] == 2]

    # Filtrar df_pedidod por LISTA_PRE igual a 2 o 'C' en df_filtrado1
    df_pedidod_filtrado1 = df_pedidod[
        (df_pedidod['NO_PED'].isin(df_filtrado1['NO_PED'])) &
        ((df_pedidod['LISTA_PRE'] == 2) | (df_pedidod['LISTA_PRE'] == 'C'))
    ]

    # Filtrar df_pedidod por LISTA_PRE igual a 1 o 'C' en df_filtrado2
    df_pedidod_filtrado2 = df_pedidod[
        (df_pedidod['NO_PED'].isin(df_filtrado2['NO_PED'])) &
        (df_clientes_filtrado_local['CVE_CTE'].isin(df_filtrado2['CVE_CTE'])) &
        ((df_pedidod['LISTA_PRE'] == 1) | (df_pedidod['LISTA_PRE'] == 'C'))
    ]
    # Filtrar df_pedidod por LISTA_PRE igual a 1 o 'C' en df_filtrado2
    df_pedidod_filtrado3 = df_pedidod[
        (df_pedidod['NO_PED'].isin(df_filtrado3['NO_PED'])) &
        (df_clientes_filtrado['CVE_CTE'].isin(df_filtrado3['CVE_CTE'])) &
        ((df_pedidod['LISTA_PRE'] == 1) | (df_pedidod['LISTA_PRE'] == 'C'))
    ]

    # Obtener el nombre de la hoja
    if ciudad == 'LAGUNA':
        sheet_name = 'LAGUNA'
    elif ciudad == 'MTY':
        sheet_name = 'MONTERREY'

    # Crear una hoja para cada ciudad
    sheet_ciudad = workbook.create_sheet(sheet_name)
    for r in dataframe_to_rows(df_pedidod_filtrado1, index=False, header=True):
        sheet_ciudad.append(r)
    for r in dataframe_to_rows(df_pedidod_filtrado2, index=False, header=True):
        sheet_ciudad.append(r)
    for r in dataframe_to_rows(df_pedidod_filtrado3, index=False, header=True):
        sheet_ciudad.append(r)

    # Crear una lista para los NO_PED agrupados
    no_ped_agregados = []

    # Agregar los NO_PED de df_pedidod_filtrado1 y df_pedidod_filtrado2 a la lista
    no_ped_agregados.extend(df_pedidod_filtrado1['NO_PED'].tolist())
    no_ped_agregados.extend(df_pedidod_filtrado2['NO_PED'].tolist())
    no_ped_agregados.extend(df_pedidod_filtrado3['NO_PED'].tolist())

    # Convertir la lista en un DataFrame y eliminar duplicados
    df_agregado_ciudad = pd.DataFrame({'CIUDAD': [ciudad] * len(no_ped_agregados), 'NO_PED': no_ped_agregados})
    df_agregado_ciudad = df_agregado_ciudad.drop_duplicates()

    # Ordenar el DataFrame por NO_PED
    df_agregado_ciudad = df_agregado_ciudad.sort_values(by=['NO_PED']).reset_index(drop=True)

    # Agregar los NO_PED agrupados a la hoja correspondiente
    for _, row in df_agregado_ciudad.iterrows():
        sheet_agregado.append([row['CIUDAD'], row['NO_PED']])

# Guardar el archivo de Excel
workbook.save(archivo_excel)
