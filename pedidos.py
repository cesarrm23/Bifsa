import pandas as pd
from datetime import datetime
from dbfread import DBF
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Ruta a los archivos DBF
pedidoc_files = {
    'LAGUNA': 'LAGUNA/pedidoc.DBF',
    'MTY': 'MTY/pedidoc.DBF'
}

pedidod_files = {
    'LAGUNA': 'LAGUNA/PEDIDOD.DBF',
    'MTY': 'MTY/PEDIDOD.DBF'
}

# Definir fecha de inicio y fecha de fin
fecha_inicio = datetime(2023, 6, 19)
fecha_fin = datetime(2023, 6, 24)

# Crear el archivo de Excel
archivo_excel = f'results/MTY_LAGUNA_del_{fecha_inicio.strftime("%Y-%m-%d")}_al_{fecha_fin.strftime("%Y-%m-%d")}.xlsx'
workbook = Workbook()

for ciudad in pedidoc_files.keys():
    # Leer el archivo pedidoc DBF
    pedidoc_dbf = DBF(pedidoc_files[ciudad])

    # Crear DataFrame de Pandas a partir de pedidoc
    df_pedidoc = pd.DataFrame(iter(pedidoc_dbf))

    # Convierte la columna de fecha en formato datetime
    df_pedidoc['F_ALTA_PED'] = pd.to_datetime(df_pedidoc['F_ALTA_PED'])

    # Filtra los registros de la semana del 22 de mayo al 27 de mayo de 2023 y con TOTAL_PED menor a 5000 en pedidoc
    df_pedidoc_filtrado = df_pedidoc[
        (df_pedidoc['F_ALTA_PED'] >= fecha_inicio) &
        (df_pedidoc['F_ALTA_PED'] <= fecha_fin) &
        (df_pedidoc['TOTAL_PED'] < 5800) &
        ((df_pedidoc['STATUS'] == 'Surtido') | (df_pedidoc['STATUS'] == 'Por Surtir'))
    ]

    # Obtener los valores únicos de NO_PED
    no_ped_values = df_pedidoc_filtrado['NO_PED'].unique()

    # Leer el archivo pedidod DBF y filtrar por NO_PED
    pedidod_dbf = DBF(pedidod_files[ciudad])
    df_pedidod = pd.DataFrame(iter(pedidod_dbf))
    df_pedidod_filtrado = df_pedidod[
        (df_pedidod['NO_PED'].isin(no_ped_values)) &
        ((df_pedidod['LISTA_PRE'] == 1) | (df_pedidod['LISTA_PRE'] == 'C'))
    ]

    # Obtener el nombre de la hoja
    if ciudad == 'LAGUNA':
        sheet_name = 'LAGUNA'
    elif ciudad == 'MTY':
        sheet_name = 'MONTERREY'

    # Agregar los datos al archivo de Excel en una nueva hoja
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)

    workbook.create_sheet(sheet_name)
    sheet = workbook[sheet_name]
    for r in dataframe_to_rows(df_pedidod_filtrado, index=False, header=True):
        sheet.append(r)

# Guardar el archivo de Excel
workbook.save(archivo_excel)
